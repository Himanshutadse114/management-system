#!/usr/bin/env python3
import argparse
import json
import math
from datetime import datetime, date
from html import escape
from pathlib import Path

TRANSLATIONS = {
    'en': {
        'generated_report': 'Management Report', 'generated_on': 'Generated on', 'period': 'Period', 'scope': 'Scope', 'all_branches': 'All branches',
        'sales': 'Sales', 'paid_orders': 'Paid orders', 'cogs': 'Cost of goods sold', 'gross_profit': 'Gross profit', 'expenses': 'Operating expenses',
        'operating_profit': 'Operating P&L', 'inventory_value': 'Inventory value', 'unresolved_orders': 'Unresolved orders', 'no_data': 'No data for this section.',
        'sales_detail': 'Sales detail', 'product_performance': 'Product & brand performance', 'alcohol_ml': 'Alcohol ML consumption',
        'inventory_valuation': 'Inventory valuation', 'stock_movements': 'Stock movements', 'wastage': 'Wastage & spillage', 'purchases': 'Purchases',
        'waiter_performance': 'Waiter reconciliation', 'unresolved_orders_section': 'Unresolved restaurant orders', 'expenses_section': 'Operating expenses',
        'branch_comparison': 'Branch comparison', 'payment_mix': 'Payment mix',
        'DAILY_CLOSING': 'Daily Closing Report', 'CONSOLIDATED': 'Consolidated Management Report', 'SALES': 'Sales Detail Report',
        'PRODUCT_PERFORMANCE': 'Product & Brand Performance', 'ALCOHOL_ML': 'Alcohol ML Consumption', 'INVENTORY_VALUATION': 'Inventory Valuation',
        'STOCK_MOVEMENTS': 'Stock Movement Ledger', 'PURCHASES': 'Purchase Report', 'WASTAGE': 'Wastage & Spillage Report',
        'WAITER_RECONCILIATION': 'Waiter Reconciliation', 'PROFIT_MARGIN': 'Profit & Margin Report', 'BRANCH_COMPARISON': 'Branch Comparison',
        'orderNumber':'Order No.','branch':'Branch','branchCode':'Branch code','orderType':'Order type','staff':'Staff','payment':'Payment','subtotalMinor':'Subtotal',
        'discountMinor':'Discount','taxMinor':'Tax','totalMinor':'Total','cogsMinor':'COGS','grossProfitMinor':'Gross profit','paidAt':'Paid at','product':'Product',
        'brand':'Brand','productType':'Type','quantityUnits':'Qty','baseQuantity':'Base qty','unit':'Unit','salesMinor':'Sales','portion':'Portion','mode':'Mode','mlSold':'ML sold',
        'quantityBase':'Stock qty','unitCostMinor':'Weighted cost / unit','inventoryValueMinor':'Inventory value','date':'Date','movementType':'Movement','quantity':'Quantity',
        'costMinor':'Cost impact','stockAfter':'Stock after','reason':'Reason','actor':'User','supplier':'Supplier','invoice':'Invoice','packageCount':'Packages','packageSize':'Package size',
        'totalBaseUnits':'Total base qty','lineTotalMinor':'Line total','name':'Name','email':'Email','paidOrders':'Paid','unresolvedOrders':'Unresolved','table':'Table','waiter':'Waiter',
        'status':'Status','openedAt':'Opened at','category':'Category','description':'Description','amountMinor':'Amount','code':'Code','type':'Type','paymentCount':'Payments','method':'Method'
    },
    'hi': {
        'generated_report': 'प्रबंधन रिपोर्ट', 'generated_on': 'तैयार किया गया', 'period': 'अवधि', 'scope': 'दायरा', 'all_branches': 'सभी शाखाएँ',
        'sales': 'बिक्री', 'paid_orders': 'भुगतान किए गए ऑर्डर', 'cogs': 'बेची गई वस्तुओं की लागत', 'gross_profit': 'सकल लाभ', 'expenses': 'परिचालन खर्च',
        'operating_profit': 'परिचालन लाभ/हानि', 'inventory_value': 'इन्वेंटरी मूल्य', 'unresolved_orders': 'अनसुलझे ऑर्डर', 'no_data': 'इस अनुभाग के लिए कोई डेटा नहीं है।',
        'sales_detail': 'बिक्री विवरण', 'product_performance': 'उत्पाद और ब्रांड प्रदर्शन', 'alcohol_ml': 'अल्कोहल ML खपत',
        'inventory_valuation': 'इन्वेंटरी मूल्यांकन', 'stock_movements': 'स्टॉक गतिविधियाँ', 'wastage': 'बर्बादी और स्पिलेज', 'purchases': 'खरीद',
        'waiter_performance': 'वेटर मिलान', 'unresolved_orders_section': 'अनसुलझे रेस्टोरेंट ऑर्डर', 'expenses_section': 'परिचालन खर्च',
        'branch_comparison': 'शाखा तुलना', 'payment_mix': 'भुगतान मिश्रण',
        'DAILY_CLOSING':'दैनिक समापन रिपोर्ट','CONSOLIDATED':'समेकित प्रबंधन रिपोर्ट','SALES':'बिक्री विवरण रिपोर्ट','PRODUCT_PERFORMANCE':'उत्पाद और ब्रांड प्रदर्शन',
        'ALCOHOL_ML':'अल्कोहल ML खपत','INVENTORY_VALUATION':'इन्वेंटरी मूल्यांकन','STOCK_MOVEMENTS':'स्टॉक मूवमेंट लेजर','PURCHASES':'खरीद रिपोर्ट',
        'WASTAGE':'बर्बादी और स्पिलेज रिपोर्ट','WAITER_RECONCILIATION':'वेटर मिलान','PROFIT_MARGIN':'लाभ और मार्जिन रिपोर्ट','BRANCH_COMPARISON':'शाखा तुलना',
        'orderNumber':'ऑर्डर नं.','branch':'शाखा','branchCode':'शाखा कोड','orderType':'ऑर्डर प्रकार','staff':'स्टाफ','payment':'भुगतान','subtotalMinor':'उप-योग','discountMinor':'छूट',
        'taxMinor':'कर','totalMinor':'कुल','cogsMinor':'लागत','grossProfitMinor':'सकल लाभ','paidAt':'भुगतान समय','product':'उत्पाद','brand':'ब्रांड','productType':'प्रकार','quantityUnits':'मात्रा',
        'baseQuantity':'आधार मात्रा','unit':'इकाई','salesMinor':'बिक्री','portion':'पोर्शन','mode':'मोड','mlSold':'ML बिक्री','quantityBase':'स्टॉक मात्रा','unitCostMinor':'औसत लागत/इकाई',
        'inventoryValueMinor':'इन्वेंटरी मूल्य','date':'तारीख','movementType':'गतिविधि','quantity':'मात्रा','costMinor':'लागत प्रभाव','stockAfter':'बाद का स्टॉक','reason':'कारण','actor':'उपयोगकर्ता',
        'supplier':'आपूर्तिकर्ता','invoice':'इनवॉइस','packageCount':'पैकेज','packageSize':'पैकेज आकार','totalBaseUnits':'कुल आधार मात्रा','lineTotalMinor':'लाइन कुल','name':'नाम','email':'ईमेल',
        'paidOrders':'भुगतान','unresolvedOrders':'अनसुलझे','table':'टेबल','waiter':'वेटर','status':'स्थिति','openedAt':'खोलने का समय','category':'श्रेणी','description':'विवरण','amountMinor':'राशि','code':'कोड','type':'प्रकार','paymentCount':'भुगतान','method':'माध्यम'
    },
    'mr': {
        'generated_report': 'व्यवस्थापन अहवाल', 'generated_on': 'तयार केले', 'period': 'कालावधी', 'scope': 'व्याप्ती', 'all_branches': 'सर्व शाखा',
        'sales': 'विक्री', 'paid_orders': 'पेड ऑर्डर्स', 'cogs': 'विक्री मालाची किंमत', 'gross_profit': 'एकूण नफा', 'expenses': 'ऑपरेटिंग खर्च',
        'operating_profit': 'ऑपरेटिंग नफा/तोटा', 'inventory_value': 'इन्व्हेंटरी मूल्य', 'unresolved_orders': 'प्रलंबित ऑर्डर्स', 'no_data': 'या विभागासाठी डेटा उपलब्ध नाही.',
        'sales_detail': 'विक्री तपशील', 'product_performance': 'उत्पादन व ब्रँड कामगिरी', 'alcohol_ml': 'अल्कोहोल ML वापर',
        'inventory_valuation': 'इन्व्हेंटरी मूल्यांकन', 'stock_movements': 'स्टॉक हालचाली', 'wastage': 'वेस्टेज व स्पिलेज', 'purchases': 'खरेदी',
        'waiter_performance': 'वेटर ताळमेळ', 'unresolved_orders_section': 'प्रलंबित रेस्टॉरंट ऑर्डर्स', 'expenses_section': 'ऑपरेटिंग खर्च',
        'branch_comparison': 'शाखा तुलना', 'payment_mix': 'पेमेंट मिश्रण',
        'DAILY_CLOSING':'दैनिक क्लोजिंग अहवाल','CONSOLIDATED':'एकत्रित व्यवस्थापन अहवाल','SALES':'विक्री तपशील अहवाल','PRODUCT_PERFORMANCE':'उत्पादन व ब्रँड कामगिरी',
        'ALCOHOL_ML':'अल्कोहोल ML वापर','INVENTORY_VALUATION':'इन्व्हेंटरी मूल्यांकन','STOCK_MOVEMENTS':'स्टॉक मूव्हमेंट लेजर','PURCHASES':'खरेदी अहवाल',
        'WASTAGE':'वेस्टेज व स्पिलेज अहवाल','WAITER_RECONCILIATION':'वेटर ताळमेळ','PROFIT_MARGIN':'नफा व मार्जिन अहवाल','BRANCH_COMPARISON':'शाखा तुलना',
        'orderNumber':'ऑर्डर क्र.','branch':'शाखा','branchCode':'शाखा कोड','orderType':'ऑर्डर प्रकार','staff':'कर्मचारी','payment':'पेमेंट','subtotalMinor':'उपएकूण','discountMinor':'सवलत',
        'taxMinor':'कर','totalMinor':'एकूण','cogsMinor':'COGS','grossProfitMinor':'एकूण नफा','paidAt':'पेमेंट वेळ','product':'उत्पादन','brand':'ब्रँड','productType':'प्रकार','quantityUnits':'नग',
        'baseQuantity':'मूळ प्रमाण','unit':'युनिट','salesMinor':'विक्री','portion':'पोर्शन','mode':'मोड','mlSold':'ML विक्री','quantityBase':'स्टॉक प्रमाण','unitCostMinor':'सरासरी किंमत/युनिट',
        'inventoryValueMinor':'इन्व्हेंटरी मूल्य','date':'तारीख','movementType':'हालचाल','quantity':'प्रमाण','costMinor':'किंमत परिणाम','stockAfter':'नंतरचा स्टॉक','reason':'कारण','actor':'वापरकर्ता',
        'supplier':'पुरवठादार','invoice':'इनव्हॉइस','packageCount':'पॅकेज','packageSize':'पॅकेज आकार','totalBaseUnits':'एकूण मूळ प्रमाण','lineTotalMinor':'लाइन एकूण','name':'नाव','email':'ईमेल',
        'paidOrders':'पेड','unresolvedOrders':'प्रलंबित','table':'टेबल','waiter':'वेटर','status':'स्थिती','openedAt':'सुरू वेळ','category':'श्रेणी','description':'वर्णन','amountMinor':'रक्कम','code':'कोड','type':'प्रकार','paymentCount':'पेमेंट्स','method':'माध्यम'
    }
}

SECTION_ALIASES = {'unresolved_orders': 'unresolved_orders_section', 'expenses': 'expenses_section'}


def t(locale, key):
    table = TRANSLATIONS.get(locale, TRANSLATIONS['en'])
    return table.get(key, TRANSLATIONS['en'].get(key, key.replace('_', ' ').title()))


def money(minor):
    try:
        n = int(str(minor or 0))
    except Exception:
        n = 0
    sign = '-' if n < 0 else ''
    n = abs(n)
    return f"{sign}₹{n / 100:,.2f}"


def decimal_value(value):
    try:
        return f"{float(value):,.3f}".rstrip('0').rstrip('.')
    except Exception:
        return str(value or '')


def display_value(value, value_type):
    if value is None:
        return '-'
    if value_type == 'money':
        return money(value)
    if value_type == 'money_per_unit':
        try:
            return f"₹{float(value) / 100:,.4f}"
        except Exception:
            return '₹0.0000'
    if value_type in ('decimal', 'number'):
        return decimal_value(value)
    if value_type in ('date', 'datetime'):
        text = str(value)
        try:
            parsed = datetime.fromisoformat(text.replace('Z', '+00:00'))
            return parsed.strftime('%d %b %Y %H:%M') if value_type == 'datetime' else parsed.strftime('%d %b %Y')
        except Exception:
            return text[:19].replace('T', ' ')
    return str(value)


def report_title(payload):
    return t(payload.get('locale', 'en'), payload.get('reportType', 'generated_report'))


def html_report(payload):
    locale = payload.get('locale', 'en')
    title = report_title(payload)
    tenant = payload.get('tenant') or {}
    branch = payload.get('branch')
    scope = branch.get('name') if branch else t(locale, 'all_branches')
    report_range = payload.get('range') or {}
    generated = str(payload.get('generatedAt') or '')[:19].replace('T', ' ')

    summary_html = ''.join(
        f'<div class="metric"><span>{escape(t(locale, item.get("key", "")))}</span><strong>{escape(display_value(item.get("value"), item.get("type")))}</strong></div>'
        for item in payload.get('summary', [])
    )

    sections_html = []
    for sec in payload.get('sections', []):
        key = sec.get('key', '')
        heading_key = SECTION_ALIASES.get(key, key)
        columns = sec.get('columns', [])
        rows = sec.get('rows', [])
        heads = ''.join(f'<th>{escape(t(locale, c.get("key", "")))}</th>' for c in columns)
        if rows:
            body = ''.join('<tr>' + ''.join(
                f'<td>{escape(display_value(row.get(c.get("key")), c.get("type", "text")))}</td>' for c in columns
            ) + '</tr>' for row in rows)
            table = f'<div class="table-wrap"><table><thead><tr>{heads}</tr></thead><tbody>{body}</tbody></table></div>'
        else:
            table = f'<div class="empty">{escape(t(locale, "no_data"))}</div>'
        sections_html.append(f'<section><h2>{escape(t(locale, heading_key))}</h2>{table}</section>')

    return f'''<!doctype html><html lang="{escape(locale)}"><head><meta charset="utf-8"><style>
      @page {{ size: A4 landscape; margin: 15mm 11mm 16mm; @bottom-center {{ content: "{escape(title)} · " counter(page) " / " counter(pages); font-size:8px; color:#7b7b7b; }} }}
      *{{box-sizing:border-box}} body{{font-family:"Noto Sans","Noto Sans Devanagari",sans-serif;color:#20242a;font-size:9px;line-height:1.45;margin:0}}
      .hero{{background:#121a19;color:#fff;border-top:5px solid #f58220;padding:20px 22px;border-radius:7px;margin-bottom:12px}}
      .kicker{{font-size:8px;font-weight:700;color:#ffaa5a;letter-spacing:.08em;text-transform:uppercase}} h1{{font-size:25px;line-height:1.05;margin:5px 0 8px}}
      .meta{{display:flex;gap:18px;flex-wrap:wrap;color:#c8cfcd;font-size:8.5px}} .meta b{{color:#fff}}
      .metrics{{display:grid;grid-template-columns:repeat(4,1fr);gap:7px;margin:10px 0 15px}} .metric{{border:1px solid #e3e3e3;border-radius:7px;padding:9px;background:#fff}}
      .metric span{{display:block;color:#73777d;font-size:7.5px;margin-bottom:3px}} .metric strong{{font-size:13px;color:#25292e}}
      section{{margin:0 0 15px;break-inside:auto}} section h2{{font-size:14px;margin:0 0 6px;border-left:3px solid #f58220;padding-left:7px}}
      .table-wrap{{border:1px solid #ddd;border-radius:6px;overflow:hidden}} table{{width:100%;border-collapse:collapse;table-layout:auto}} thead{{display:table-header-group}}
      th{{background:#f2eee9;color:#5a524c;font-size:7.3px;text-align:left;padding:6px 5px;border-bottom:1px solid #ddd;white-space:nowrap}}
      td{{padding:5px;border-bottom:1px solid #eee;vertical-align:top;font-size:7.1px;overflow-wrap:anywhere}} tbody tr:nth-child(even){{background:#fbfaf8}} tr{{break-inside:avoid}}
      .empty{{padding:18px;border:1px dashed #d6d0ca;border-radius:6px;color:#777;text-align:center}} .footer-note{{margin-top:8px;color:#888;font-size:7.5px}}
    </style></head><body>
      <div class="hero"><div class="kicker">OUTLET OS · {escape(t(locale,'generated_report'))}</div><h1>{escape(title)}</h1>
      <div class="meta"><span><b>{escape(tenant.get('name',''))}</b></span><span>{escape(t(locale,'scope'))}: <b>{escape(scope)}</b></span><span>{escape(t(locale,'period'))}: <b>{escape(report_range.get('from',''))} → {escape(report_range.get('to',''))}</b></span><span>{escape(t(locale,'generated_on'))}: {escape(generated)}</span></div></div>
      <div class="metrics">{summary_html}</div>{''.join(sections_html)}
    </body></html>'''


def generate_pdf(payload, output_path):
    from weasyprint import HTML
    HTML(string=html_report(payload), base_url=str(Path(output_path).parent)).write_pdf(output_path)


def excel_number(value, value_type):
    if value is None:
        return None
    try:
        if value_type in ('money', 'money_per_unit'):
            return float(value) / 100.0
        if value_type in ('decimal', 'number'):
            return float(value)
    except Exception:
        pass
    return str(value)


def generate_xlsx(payload, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    locale = payload.get('locale', 'en')
    font_name = 'Noto Sans Devanagari' if locale in ('hi', 'mr') else 'Noto Sans'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    orange = 'F58220'; dark = '121A19'; cream = 'F3EEE9'; white = 'FFFFFF'; thin = Side(style='thin', color='E3E3E3')
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:D1'); ws['A1'] = report_title(payload); ws['A1'].font = Font(name=font_name, bold=True, size=18, color=white); ws['A1'].fill = PatternFill('solid', fgColor=dark); ws['A1'].alignment = Alignment(vertical='center'); ws.row_dimensions[1].height = 30
    meta = [
        (t(locale,'scope'), (payload.get('branch') or {}).get('name') or t(locale,'all_branches')),
        (t(locale,'period'), f"{payload.get('range',{}).get('from','')} → {payload.get('range',{}).get('to','')}"),
        (t(locale,'generated_on'), str(payload.get('generatedAt',''))[:19].replace('T',' '))
    ]
    row = 3
    for label, value in meta:
        ws.cell(row,1,label).font=Font(name=font_name,bold=True,color='6B625B'); ws.cell(row,2,value).font=Font(name=font_name); row += 1
    row += 1
    for item in payload.get('summary', []):
        ws.cell(row,1,t(locale,item.get('key',''))).font=Font(name=font_name,bold=True)
        cell=ws.cell(row,2,excel_number(item.get('value'),item.get('type'))); cell.font=Font(name=font_name,bold=True,color='A84600')
        if item.get('type')=='money': cell.number_format='₹#,##0.00;[Red]-₹#,##0.00'
        row += 1
    ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=32

    used_names={'Summary'}
    for idx, sec in enumerate(payload.get('sections', []), start=1):
        base=t(locale,SECTION_ALIASES.get(sec.get('key',''),sec.get('key','')))[:27] or f'Section {idx}'
        name=base; suffix=1
        while name in used_names:
            suffix+=1; name=f'{base[:24]} {suffix}'
        used_names.add(name)
        sh=wb.create_sheet(name); sh.sheet_view.showGridLines=False
        columns=sec.get('columns',[]); rows=sec.get('rows',[])
        for c_idx,c in enumerate(columns,1):
            cell=sh.cell(1,c_idx,t(locale,c.get('key',''))); cell.font=Font(name=font_name,bold=True,color=white); cell.fill=PatternFill('solid',fgColor=dark); cell.alignment=Alignment(vertical='center'); cell.border=Border(bottom=thin)
        sh.freeze_panes='A2'; sh.auto_filter.ref=f'A1:{get_column_letter(max(1,len(columns)))}{max(1,len(rows)+1)}'
        widths=[len(t(locale,c.get('key','')))+2 for c in columns]
        for r_idx,data in enumerate(rows,2):
            for c_idx,c in enumerate(columns,1):
                raw=data.get(c.get('key')); typ=c.get('type','text'); val=excel_number(raw,typ); cell=sh.cell(r_idx,c_idx,val); cell.font=Font(name=font_name,size=9); cell.alignment=Alignment(vertical='top',wrap_text=True); cell.border=Border(bottom=Side(style='hair',color='EEEEEE'))
                if typ=='money': cell.number_format='₹#,##0.00;[Red]-₹#,##0.00'
                elif typ=='money_per_unit': cell.number_format='₹#,##0.0000;[Red]-₹#,##0.0000'
                elif typ in ('decimal','number'): cell.number_format='#,##0.000'
                widths[c_idx-1]=min(42,max(widths[c_idx-1],len(str(val or ''))+2))
        for c_idx,width in enumerate(widths,1): sh.column_dimensions[get_column_letter(c_idx)].width=max(10,width)
        sh.row_dimensions[1].height=23
    wb.save(output_path)


def main():
    parser=argparse.ArgumentParser()
    parser.add_argument('--input',required=True);parser.add_argument('--output',required=True);parser.add_argument('--format',choices=['pdf','xlsx'],required=True)
    args=parser.parse_args()
    with open(args.input,'r',encoding='utf-8') as fh: payload=json.load(fh)
    Path(args.output).parent.mkdir(parents=True,exist_ok=True)
    if args.format=='pdf': generate_pdf(payload,args.output)
    else: generate_xlsx(payload,args.output)
    print(json.dumps({'ok':True,'output':args.output},ensure_ascii=False))

if __name__=='__main__':
    main()
